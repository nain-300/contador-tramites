#!/usr/bin/env python3
"""
Contador de Trámites — Aplicación de escritorio para contar trámites diarios.
Hotkey global configurable (F1 por defecto) para sumar un trámite.
"""

import json
import os
import sys
import threading
from datetime import datetime, date
import tkinter as tk
from tkinter import messagebox
import platform

import keyboard

# ─── Versión ─────────────────────────────────────────────────────────────────
VERSION = "1.0.3"

# ─── Rutas ───────────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(SCRIPT_DIR, "data.json")
EXCEL_FILE = os.path.join(SCRIPT_DIR, "historial_tramites.xlsx")

# ─── Colores y estilos ──────────────────────────────────────────────────────
BG_COLOR = "#1e1e2e"
BG_LIGHTER = "#2a2a3d"
TEXT_PRIMARY = "#a6e3a1"       # verde claro para el número
TEXT_SECONDARY = "#cdd6f4"     # texto claro
TEXT_MUTED = "#6c7086"         # texto apagado
ACCENT = "#f38ba8"             # rosa para botón reset
ACCENT_HOVER = "#f5a0b8"
BORDER_COLOR = "#45475a"
BLUE_ACCENT = "#89b4fa"        # azul para botón hotkey
BLUE_HOVER = "#a6c8ff"
GREEN_ACCENT = "#a6e3a1"       # verde para botón exportar
GREEN_HOVER = "#b8f0b0"
RED_ACCENT = "#ff5555"         # rojo para botón deshacer
RED_HOVER = "#ff7777"
RED_MUTED = "#8b3a3a"          # rojo oscuro deshabilitado

DEFAULT_HOTKEY = "F1"


# ─── Funciones de persistencia ───────────────────────────────────────────────

def load_data() -> dict:
    """Carga los datos del archivo JSON. Si es un día nuevo, reinicia contadores
    pero conserva la hotkey configurada."""
    today = date.today().isoformat()
    default = {"date": today, "count": 0, "timestamps": [], "hotkey": DEFAULT_HOTKEY}

    if not os.path.exists(DATA_FILE):
        return default

    try:
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
    except (json.JSONDecodeError, IOError):
        return default

    # Preservar hotkey configurada
    saved_hotkey = data.get("hotkey", DEFAULT_HOTKEY)

    # Si la fecha guardada no es hoy, empezar de cero pero conservar hotkey
    if data.get("date") != today:
        default["hotkey"] = saved_hotkey
        return default

    # Asegurar que el campo hotkey exista
    data.setdefault("hotkey", DEFAULT_HOTKEY)
    return data


def save_data(data: dict) -> None:
    """Guarda los datos en el archivo JSON."""
    try:
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except IOError as e:
        print(f"Error al guardar datos: {e}", file=sys.stderr)


# ─── Aplicación principal ───────────────────────────────────────────────────

class ContadorTramitesApp:
    """Ventana principal del contador de trámites."""

    def __init__(self):
        self.data = load_data()
        self._hotkey_handle = None  # referencia al hotkey registrado

        # ── Ventana ──────────────────────────────────────────────────────
        self.root = tk.Tk()
        
        # Cargar icono
        icon_path = os.path.join(SCRIPT_DIR, "icono.ico")
        if os.path.exists(icon_path):
            try:
                if platform.system() == "Windows":
                    self.root.iconbitmap(icon_path)
                else:
                    icon_img = tk.PhotoImage(file=icon_path)
                    self.root.iconphoto(True, icon_img)
            except Exception as e:
                print(f"Error cargando icono: {e}", file=sys.stderr)
                
        self.root.title("Contador")
        self.root.overrideredirect(True)          # Sin barra de título
        self.root.attributes("-topmost", True)     # Siempre encima
        self.root.configure(bg=BG_COLOR)
        self.root.resizable(False, False)

        # Tamaño y posición: esquina superior derecha
        self._win_w, self._win_h = 220, 160
        screen_w = self.root.winfo_screenwidth()
        x = screen_w - self._win_w - 16
        y = 16
        self._win_pos = (x, y)
        self.root.geometry(f"{self._win_w}x{self._win_h}+{x}+{y}")

        # ── Drag para mover la ventana ───────────────────────────────────
        self._drag_x = 0
        self._drag_y = 0
        self.root.bind("<ButtonPress-1>", self._start_drag)
        self.root.bind("<B1-Motion>", self._on_drag)

        # ── Construir UI ─────────────────────────────────────────────────
        self._build_ui()

        # ── Actualizar valores iniciales ─────────────────────────────────
        self._refresh_display()
        save_data(self.data)  # Guardar por si cambió la fecha al cargar

        # ── Check automático de nuevo día ────────────────────────────────
        self._check_new_day()

        # ── Hotkey global en hilo aparte ─────────────────────────────────
        self._register_hotkey()

        # ── Cerrar limpiamente ───────────────────────────────────────────
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    # ── UI ────────────────────────────────────────────────────────────────

    def _build_ui(self):
        # Barra superior mínima (título + cerrar)
        top_bar = tk.Frame(self.root, bg=BG_LIGHTER, height=24)
        top_bar.pack(fill="x")
        top_bar.pack_propagate(False)

        title_lbl = tk.Label(
            top_bar, text="⚡ Trámites", font=("Segoe UI", 9),
            bg=BG_LIGHTER, fg=TEXT_MUTED, anchor="w", padx=6
        )
        title_lbl.pack(side="left", fill="x", expand=True)

        close_btn = tk.Label(
            top_bar, text="✕", font=("Segoe UI", 10, "bold"),
            bg=BG_LIGHTER, fg=TEXT_MUTED, padx=8, cursor="hand2"
        )
        close_btn.pack(side="right")
        close_btn.bind("<Button-1>", lambda e: self._on_close())
        close_btn.bind("<Enter>", lambda e: close_btn.config(fg=ACCENT))
        close_btn.bind("<Leave>", lambda e: close_btn.config(fg=TEXT_MUTED))

        # Contenedor central
        body = tk.Frame(self.root, bg=BG_COLOR, padx=10, pady=4)
        body.pack(fill="both", expand=True)

        # Número grande del contador
        self.count_var = tk.StringVar(value="0")
        count_lbl = tk.Label(
            body, textvariable=self.count_var,
            font=("Segoe UI", 48, "bold"),
            bg=BG_COLOR, fg=TEXT_PRIMARY
        )
        count_lbl.pack(anchor="center")

        # Línea inferior: info + reset
        bottom = tk.Frame(body, bg=BG_COLOR)
        bottom.pack(fill="x", pady=(0, 2))



        # Botón exportar Excel
        export_btn = tk.Label(
            bottom, text="⬇ Exportar", font=("Segoe UI", 9),
            bg=BG_COLOR, fg=GREEN_ACCENT, cursor="hand2", padx=4
        )
        export_btn.pack(side="right", padx=(0, 4))
        export_btn.bind("<Button-1>", lambda e: self._export_to_excel(show_msg=True))
        export_btn.bind("<Enter>", lambda e: export_btn.config(fg=GREEN_HOVER))
        export_btn.bind("<Leave>", lambda e: export_btn.config(fg=GREEN_ACCENT))

        # Botón deshacer
        self.undo_btn = tk.Label(
            bottom, text="Deshacer", font=("Segoe UI", 9),
            bg=BG_COLOR, fg=RED_MUTED, padx=4, width=8
        )
        self.undo_btn.pack(side="right", padx=(0, 2))
        self.undo_btn.bind("<Button-1>", lambda e: self._undo())

        def on_enter_undo(e):
            if self.data["count"] > 0:
                self.undo_btn.config(fg=RED_HOVER)

        def on_leave_undo(e):
            if self.data["count"] > 0:
                self.undo_btn.config(fg=RED_ACCENT)

        self.undo_btn.bind("<Enter>", on_enter_undo)
        self.undo_btn.bind("<Leave>", on_leave_undo)

        # Línea de hotkey + último trámite
        hotkey_row = tk.Frame(body, bg=BG_COLOR)
        hotkey_row.pack(fill="x", pady=(0, 2))

        # Botón para cambiar hotkey
        self.hotkey_var = tk.StringVar(value=f"⌨ Tecla: {self.data['hotkey']}")
        self.hotkey_btn = tk.Label(
            hotkey_row, textvariable=self.hotkey_var,
            font=("Segoe UI", 8), bg=BG_COLOR, fg=BLUE_ACCENT,
            cursor="hand2"
        )
        self.hotkey_btn.pack(side="left")
        self.hotkey_btn.bind("<Button-1>", lambda e: self._open_hotkey_config())
        self.hotkey_btn.bind("<Enter>", lambda e: self.hotkey_btn.config(fg=BLUE_HOVER))
        self.hotkey_btn.bind("<Leave>", lambda e: self.hotkey_btn.config(fg=BLUE_ACCENT))

        # Hora del último trámite
        self.last_var = tk.StringVar(value="")
        last_lbl = tk.Label(
            hotkey_row, textvariable=self.last_var,
            font=("Segoe UI", 8), bg=BG_COLOR, fg=TEXT_MUTED
        )
        last_lbl.pack(side="right")

    # ── Lógica ────────────────────────────────────────────────────────────

    def _increment(self):
        """Suma 1 al contador, guarda timestamp y persiste."""
        now = datetime.now()
        self.data["count"] += 1
        self.data["timestamps"].append(now.strftime("%H:%M:%S"))
        save_data(self.data)
        # Actualizar UI desde el hilo principal
        self.root.after(0, self._refresh_display)

    def _undo(self):
        """Resta 1 al contador y elimina el último timestamp."""
        if self.data["count"] <= 0:
            return
        self.data["count"] -= 1
        if self.data["timestamps"]:
            self.data["timestamps"].pop()
        save_data(self.data)
        self._refresh_display()

    def _refresh_display(self):
        """Actualiza los labels con los valores actuales."""
        self.count_var.set(str(self.data["count"]))
        hotkey = self.data.get("hotkey", DEFAULT_HOTKEY)
        self.hotkey_var.set(f"⌨ Tecla: {hotkey}")

        if self.data["timestamps"]:
            last_ts = self.data["timestamps"][-1]
            self.last_var.set(f"Últ: {last_ts}")
        else:
            self.last_var.set("")

        # Habilitar/deshabilitar botón deshacer
        if self.data["count"] > 0:
            self.undo_btn.config(fg=RED_ACCENT, cursor="hand2")
        else:
            self.undo_btn.config(fg=RED_MUTED, cursor="")

    def _check_new_day(self):
        """Verifica periódicamente si cambió el día para resetear automáticamente."""
        today = date.today().isoformat()
        if self.data.get("date") != today:
            # Auto-exportar el historial del día anterior si tenía trámites
            if self.data.get("count", 0) > 0:
                self._export_to_excel(show_msg=False)

            self.data["date"] = today
            self.data["count"] = 0
            self.data["timestamps"] = []
            save_data(self.data)
            self._refresh_display()

        # Comprobar de nuevo en 1 minuto (60000 ms)
        self.root.after(60000, self._check_new_day)

    # ── Hotkey global ─────────────────────────────────────────────────────

    def _register_hotkey(self):
        """Registra (o re-registra) la hotkey global configurada."""
        # Desregistrar hotkey anterior si existe
        if self._hotkey_handle is not None:
            keyboard.remove_hotkey(self._hotkey_handle)
            self._hotkey_handle = None

        hotkey = self.data.get("hotkey", DEFAULT_HOTKEY)
        self._hotkey_handle = keyboard.add_hotkey(
            hotkey, self._increment, suppress=False
        )

    # ── Configuración de hotkey ───────────────────────────────────────────

    def _open_hotkey_config(self):
        """Abre una ventana secundaria para cambiar la hotkey."""
        dialog = tk.Toplevel(self.root)
        dialog.title("Configurar tecla")
        dialog.configure(bg=BG_COLOR)
        dialog.attributes("-topmost", True)
        dialog.resizable(False, False)
        dialog.geometry("260x160")

        # Centrar respecto a la ventana principal
        dialog.transient(self.root)
        dialog.update_idletasks()
        dialog.lift()
        dialog.focus_force()
        dialog.after(150, dialog.grab_set)

        current_hotkey = self.data.get("hotkey", DEFAULT_HOTKEY)
        captured_key = {"value": None}  # mutable para el thread
        capturing = {"active": False}

        # ── Título ──
        tk.Label(
            dialog, text="Configurar hotkey",
            font=("Segoe UI", 11, "bold"), bg=BG_COLOR, fg=TEXT_SECONDARY
        ).pack(pady=(10, 4))

        # ── Hotkey actual ──
        tk.Label(
            dialog, text=f"Actual: {current_hotkey}",
            font=("Segoe UI", 9), bg=BG_COLOR, fg=TEXT_MUTED
        ).pack()

        # ── Campo de captura ──
        capture_var = tk.StringVar(value="Hacé clic aquí y presioná una tecla")
        capture_lbl = tk.Label(
            dialog, textvariable=capture_var,
            font=("Segoe UI", 10), bg=BG_LIGHTER, fg=TEXT_PRIMARY,
            relief="flat", padx=10, pady=6, cursor="hand2"
        )
        capture_lbl.pack(padx=12, pady=8, fill="x")

        def start_capture(event=None):
            if capturing["active"]:
                return
            capturing["active"] = True
            capture_var.set("⏳ Presioná una tecla...")

            def read_key():
                try:
                    event = keyboard.read_event(suppress=False)
                    if event.event_type == keyboard.KEY_DOWN:
                        # Construir nombre con modificadores
                        modifiers = []
                        if keyboard.is_pressed("ctrl"):
                            modifiers.append("ctrl")
                        if keyboard.is_pressed("alt"):
                            modifiers.append("alt")
                        if keyboard.is_pressed("shift"):
                            modifiers.append("shift")

                        key_name = event.name
                        # No duplicar el modificador en el nombre
                        if key_name.lower() not in ("ctrl", "alt", "shift",
                                                     "left ctrl", "right ctrl",
                                                     "left alt", "right alt",
                                                     "left shift", "right shift"):
                            if modifiers:
                                combo = "+".join(modifiers) + "+" + key_name
                            else:
                                combo = key_name
                            captured_key["value"] = combo
                            dialog.after(0, lambda: capture_var.set(f"✅ {combo}"))
                        else:
                            # Solo presionó un modificador, seguir esperando
                            capturing["active"] = False
                            dialog.after(0, start_capture)
                            return
                except Exception:
                    dialog.after(0, lambda: capture_var.set("❌ Error al capturar"))
                finally:
                    capturing["active"] = False

            threading.Thread(target=read_key, daemon=True).start()

        capture_lbl.bind("<Button-1>", start_capture)

        # ── Botones ──
        btn_frame = tk.Frame(dialog, bg=BG_COLOR)
        btn_frame.pack(pady=(0, 10))

        def save_hotkey():
            new_key = captured_key["value"]
            if new_key:
                self.data["hotkey"] = new_key
                save_data(self.data)
                self._register_hotkey()
                self._refresh_display()
            dialog.destroy()

        def cancel():
            dialog.destroy()

        save_btn = tk.Label(
            btn_frame, text="  Guardar  ", font=("Segoe UI", 9, "bold"),
            bg=BLUE_ACCENT, fg=BG_COLOR, cursor="hand2", padx=8, pady=2
        )
        save_btn.pack(side="left", padx=4)
        save_btn.bind("<Button-1>", lambda e: save_hotkey())

        cancel_btn = tk.Label(
            btn_frame, text="  Cancelar  ", font=("Segoe UI", 9),
            bg=BG_LIGHTER, fg=TEXT_MUTED, cursor="hand2", padx=8, pady=2
        )
        cancel_btn.pack(side="left", padx=4)
        cancel_btn.bind("<Button-1>", lambda e: cancel())

    # ── Drag ──────────────────────────────────────────────────────────────

    def _start_drag(self, event):
        self._drag_x = event.x
        self._drag_y = event.y

    def _on_drag(self, event):
        x = self.root.winfo_x() + event.x - self._drag_x
        y = self.root.winfo_y() + event.y - self._drag_y
        self._win_pos = (x, y)
        self.root.geometry(f"{self._win_w}x{self._win_h}+{x}+{y}")

    # ── Cierre ────────────────────────────────────────────────────────────

    def _on_close(self):
        save_data(self.data)
        self._export_to_excel(show_msg=False)
        keyboard.unhook_all()
        self.root.destroy()

    # ── Exportar a Excel ─────────────────────────────────────────────────

    def _export_to_excel(self, show_msg=False):
        """Exporta el historial del día a un archivo Excel.
        Cada día es una hoja con formato DD-MM-YYYY."""
        try:
            from openpyxl import load_workbook
            from openpyxl import Workbook
        except ImportError:
            if show_msg:
                messagebox.showerror(
                    "Error",
                    "No se encontró openpyxl.\n"
                    "Ejecutá: pip install openpyxl",
                    parent=self.root
                )
            return

        timestamps = self.data.get("timestamps", [])
        if not timestamps:
            if show_msg:
                messagebox.showinfo(
                    "Exportar",
                    "No hay trámites para exportar hoy.",
                    parent=self.root
                )
            return

        excel_path = self.data.get("excel_path")
        if not excel_path:
            from tkinter import filedialog
            if platform.system() == "Windows":
                initial_dir = os.path.expanduser("~/Documents")
            else:
                initial_dir = os.path.expanduser("~/Documentos")
                
            excel_path = filedialog.asksaveasfilename(
                parent=self.root,
                title="Guardar Excel",
                initialdir=initial_dir,
                initialfile="historial_tramites.xlsx",
                defaultextension=".xlsx",
                filetypes=[("Archivos Excel", "*.xlsx")]
            )
            
            if not excel_path:
                return
                
            self.data["excel_path"] = excel_path
            save_data(self.data)

        # Nombre de la hoja: DD-MM-YYYY
        raw_date = self.data.get("date", date.today().isoformat())
        try:
            parts = raw_date.split("-")  # YYYY-MM-DD
            sheet_name = f"{parts[2]}-{parts[1]}-{parts[0]}"
        except (IndexError, AttributeError):
            sheet_name = raw_date

        # Abrir o crear el workbook
        if os.path.exists(excel_path):
            try:
                wb = load_workbook(excel_path)
            except Exception:
                wb = Workbook()
        else:
            wb = Workbook()
            # Eliminar la hoja por defecto "Sheet" si existe
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        # Si la hoja para hoy ya existe, reemplazarla
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

        ws = wb.create_sheet(title=sheet_name)

        from openpyxl.styles import Font
        bold_font = Font(bold=True)

        # Encabezados (fila 1)
        for col, header in [("A", "Hora"), ("B", "Cantidad"), ("C", "Acumulado")]:
            cell = ws[f"{col}1"]
            cell.value = header
            cell.font = bold_font

        # Escribir datos: cada fila es un trámite
        for i, ts in enumerate(timestamps, start=2):
            ws[f"A{i}"] = ts
            ws[f"B{i}"] = 1
            ws[f"C{i}"] = i - 1  # acumulado: 1, 2, 3...

        # Fila de totales (una fila vacía + fila de total)
        total_row = len(timestamps) + 3
        total = len(timestamps)
        ws[f"A{total_row}"] = "TOTAL"
        ws[f"A{total_row}"].font = bold_font
        ws[f"B{total_row}"] = total
        ws[f"B{total_row}"].font = bold_font
        ws[f"C{total_row}"] = total
        ws[f"C{total_row}"].font = bold_font

        # Ajustar ancho de columnas
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12

        try:
            wb.save(excel_path)
            if show_msg:
                self._show_success_popup(f"Guardado en:\n{excel_path}")
        except IOError as e:
            if show_msg:
                messagebox.showerror(
                    "Error",
                    f"No se pudo guardar el archivo:\n{e}",
                    parent=self.root
                )

    def _show_success_popup(self, msg):
        dialog = tk.Toplevel(self.root)
        dialog.overrideredirect(True)
        dialog.configure(bg=BG_LIGHTER, highlightbackground=BORDER_COLOR, highlightthickness=1)
        
        x = self.root.winfo_x()
        y = self.root.winfo_y() + self.root.winfo_height() + 5
        dialog.geometry(f"+{x}+{y}")
        
        dialog.transient(self.root)
        dialog.update_idletasks()
        dialog.lift()
        
        lbl = tk.Label(
            dialog, text="✅ Exportación Exitosa\n" + msg,
            font=("Segoe UI", 8), bg=BG_LIGHTER, fg=TEXT_SECONDARY,
            justify="left", padx=8, pady=6
        )
        lbl.pack()
        
        close_func = lambda e=None: dialog.destroy()
        dialog.bind("<Button-1>", close_func)
        lbl.bind("<Button-1>", close_func)
        dialog.after(3000, close_func)

    # ── Run ───────────────────────────────────────────────────────────────

    def run(self):
        self.root.mainloop()


def check_for_updates():
    """Busca actualizaciones en GitHub usando version.txt, reemplaza el archivo local y reinicia si es necesario."""
    version_url = "https://raw.githubusercontent.com/nain-300/contador-tramites/master/version.txt"
    code_url = "https://raw.githubusercontent.com/nain-300/contador-tramites/master/contador.py"
    try:
        import urllib.request
        
        # Consultar la versión remota
        req_version = urllib.request.Request(version_url, headers={'Cache-Control': 'no-cache'})
        with urllib.request.urlopen(req_version, timeout=3) as response:
            remote_version = response.read().decode('utf-8').strip()
            
        def parse_version(v):
            return tuple(map(int, v.strip().split(".")))
            
        # Si la versión remota es mayor a la local, se actualiza
        if parse_version(remote_version) > parse_version(VERSION):
            req_code = urllib.request.Request(code_url, headers={'Cache-Control': 'no-cache'})
            with urllib.request.urlopen(req_code, timeout=3) as response:
                remote_code = response.read().decode('utf-8').replace('\r\n', '\n')
                
            if remote_code:
                with open(os.path.abspath(__file__), 'w', encoding='utf-8') as f:
                    f.write(remote_code)
                os.execv(sys.executable, [sys.executable] + sys.argv)
    except Exception:
        if sys.stdout is not None:
            try:
                print("sin conexión o error al actualizar, usando versión local")
            except Exception:
                pass

# ─── Entry point ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    check_for_updates()
    
    if platform.system() == "Linux":
        if hasattr(os, "geteuid") and os.geteuid() != 0:
            print("ADVERTENCIA: En Linux, la librería 'keyboard' requiere permisos de superusuario.")
            print("Si la aplicación falla o no captura las teclas, ejecútala con 'sudo'.")
            
    app = ContadorTramitesApp()
    app.run()
